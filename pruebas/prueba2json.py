from code import compile_command
from ctypes import alignment
from tkinter import HORIZONTAL
import pandas as pd
import xlsxwriter
from openpyxl.workbook import Workbook
import openpyxl
from openpyxl import Workbook
import time
from openpyxl.chart import ScatterChart, Reference, Series
from recursos_excel import*
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
import sys
import os
#listas
seccion=[]
fechas=[]
descripcion=[]
total=[]
importe=[]
matriz=[]
factor_conversion=1.10

def factura_excel(datajson,tipo_moneda='S/.'):#euro va ala derecha
    total_monto=0
    book=Workbook()
    sheet= book.active
    for columna in range(1,6):
        col_letter = get_column_letter(columna)
        sheet.column_dimensions[col_letter].width = 20 #NO ESTA EN PIXELES
    c=datajson["Seccion"]

    #CREANDO LA TABLA
    #toda esta parte es fija en la tabla
    #----------------------------------------------------------------
    combinar_celdas(sheet,'B1','D1','JUNTA DE PROPIETARIOS')
    sheet['B1'].alignment=Alignment(horizontal="center")
    combinar_celdas(sheet,'B2','D2','EDIFICIO GALLITO DE LAS ROCAS')
    sheet['B2'].alignment=Alignment(horizontal="center")
    combinar_celdas(sheet,'B3','D3','Calle Las Palomas 534 - Surquillo')
    sheet['B3'].alignment=Alignment(horizontal="center")
    combinar_celdas(sheet,'B4','D4','RECIBO POR CUOTA DE MANTENIMIENTO')
    sheet['B4'].alignment=Alignment(horizontal="center")
    formato_celdas(sheet,'B1','Arial',10,True,True,'000000',True)
    combinar_celdas(sheet,'A1','A4')
    combinar_celdas(sheet,'E1','E4')

    sheet['A5']='Departamento:'
    sheet['A6']='Propietario:'
    sheet['A7']='Periodo:'

    sheet['B5']=512
    sheet['B5'].alignment=Alignment(horizontal='center')

    sheet['B7']='SETIEMBRE 2022'
    sheet['B7'].alignment=Alignment(horizontal='center')

    sheet['C5']='Estacionamiento:'
    sheet.merge_cells('B6:E6')
    sheet['B6']='Eduardo Alberto Berrios Villanueva'
    sheet['B6'].alignment=Alignment(horizontal='center')
    sheet['C7']='Total Porc. Part.'

    sheet.merge_cells('D5:E5')
    sheet['D5']=105
    sheet['D5'].alignment=Alignment(horizontal='center')
    sheet.merge_cells('D7:E7')
    sheet['D7']=factor_conversion
    sheet['D7'].alignment=Alignment(horizontal='center')

    sheet['D8']='Total'
    formato_celdas(sheet,'D8','Arial',10,False,True,'000000',True)
    sheet['D8'].alignment=Alignment(horizontal='center')

    sheet['E8']='Importe'
    formato_celdas(sheet,'E8','Arial',10,False,True,'000000',True)
    sheet['E8'].alignment=Alignment(horizontal='center')

    for fila in range(1,8):
        for columna in range(1,6):
            sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                        right=Side(border_style='thin', color='000000'),
                                                        top=Side(border_style='thin', color='000000'),
                                                        bottom=Side(border_style='thin', color='000000'))

    sheet.cell(7,6).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style=None, color='000000'),
                                    top=Side(border_style=None, color='000000'),
                                    bottom=Side(border_style=None, color='000000'))
    #----------------------------------------------------------------
    iterable=9
    #en esta seccion se llenan los datos de la BD
    #----------------------------------------------------------------
    for j in range(2):
        nombre_seccion=c[j]["ID_Seccion"]
        b=c[j]['Subsecciones']
        celda_seccion=f'A{iterable}'
        sheet[celda_seccion]=nombre_seccion
        formato_celdas(sheet,celda_seccion,'Arial',10,True,True,'000000',True)
        iterable=iterable+1
        for i in b:
            #print('ID_Subseccion: ',i['ID_Subseccion'])
            seccion.append(i['ID_Subseccion'])
            celda_col1=f'A{iterable}'
            valor_col1=i['ID_Subseccion']
            #sheet.write(celda_col1, valor_col1)
            sheet[celda_col1] = valor_col1

            #print('ID_Subseccion: ',i['nombre'])
            fechas.append(i['nombre'])
            celda_col2=f'B{iterable}'
            valor_col2=i['nombre']
            #sheet.write(celda_col2, valor_col2)
            sheet[celda_col2] = valor_col2

            #print('ID_Subseccion: ',i['descripcion'])
            descripcion.append(i['descripcion'])

            #print('ID_Subseccion: ',i['monto'])
            total.append(i['monto'])
            celda_col4=f'D{iterable}'
            monto=i['monto']
            monto_float = "{:.2f}".format(monto)
            if (tipo_moneda!='€'):
                valor_col4=tipo_moneda +' '+str(monto_float)
            else:
                valor_col4=str(monto_float)+' '+tipo_moneda 
            #sheet.write(celda_col4, valor_col4)
            sheet[celda_col4] = valor_col4
            sheet[celda_col4].alignment=Alignment(horizontal='center')

            importe.append(i['monto']*factor_conversion)
            celda_col5=f'E{iterable}'
            total_monto_fila=i['monto']*factor_conversion
            total_monto=total_monto+total_monto_fila
            total_monto_fila_float="{:.2f}".format(total_monto_fila)
            if (tipo_moneda!='€'):
                valor_col5=tipo_moneda +' '+str(total_monto_fila_float)
            else:
                valor_col5=str(total_monto_fila_float)+' '+tipo_moneda
            #sheet.write(celda_col5, valor_col5)

            sheet[celda_col5] = valor_col5
            sheet[celda_col5].alignment=Alignment(horizontal='center')

            celda_final_suma=celda_col5
            iterable = iterable + 1

    #La suma total
    total_monto_float="{:.2f}".format(total_monto)
    if (tipo_moneda!='€'):
        celda_suma_expresion=tipo_moneda +' '+str(total_monto_float)
    else:
        celda_suma_expresion=str(total_monto_float)+' '+tipo_moneda
    celda_total=f'A{iterable}'
    celda_total_valor=f'E{iterable}'
    sheet[celda_total] = 'TOTAL'
    formato_celdas(sheet,celda_total,'Arial',10,True,False,'000000')
    sheet[celda_total_valor]=celda_suma_expresion
    sheet[celda_total_valor].alignment=Alignment(horizontal='center')
    formato_celdas(sheet,celda_total_valor,'Arial',10,True,False,'000000')
    iterable= iterable + 1
    #----------------------------------------------------------------

    #ultimos datos de abajo
    #----------------------------------------------------------------
    celda_fecha_emision=f'A{iterable}'
    sheet[celda_fecha_emision]='Fecha de emision' #inicial
    
    celda_fecha_emision=f'A{iterable+1}'
    valor_fecha_emision='01/09/2022'
    sheet[celda_fecha_emision]=valor_fecha_emision

    celda_fecha_vencimento=f'B{iterable}'
    sheet[celda_fecha_vencimento]='Fecha de vencimiento'

    celda_fecha_vencimento=f'B{iterable+1}'
    valor_fecha_vencimento='07/07/2022'
    sheet[celda_fecha_vencimento]=valor_fecha_vencimento

    celda_ncuenta=f'C{iterable}'
    sheet[celda_ncuenta]='N° Cuenta'

    celda_CCI=f'C{iterable+1}'
    sheet[celda_CCI]='CCI'

    valor_ini_ncuenta=f'D{iterable}'
    ncuenta='194-123456789'
    sheet[valor_ini_ncuenta]=ncuenta
    
    valor_fin_ncuenta=f'E{iterable}'

    valor_ini_CCI=f'D{iterable+1}'
    cci="0021194132456789" #final
    sheet[valor_ini_CCI]=cci

    valor_fin_CCI=f'E{iterable+1}'
    ultima_fila=sheet.max_row

    for i in range(iterable,ultima_fila+1):
        for j in range(1,6):
            sheet.cell(i,j).border = Border(left=Side(border_style='thin', color='000000'),
                                            right=Side(border_style='thin', color='000000'),
                                            top=Side(border_style='thin', color='000000'),
                                            bottom=Side(border_style='thin', color='000000'))
            letra=get_column_letter(j)
            celda=f'{letra}{i}'
            sheet[celda].alignment=Alignment(horizontal='center')
    combinar_celdas(sheet,valor_ini_ncuenta,valor_fin_ncuenta)
    combinar_celdas(sheet,valor_ini_CCI,valor_fin_CCI)
    #estilos de las celdas (bordes)
    #'mediumDashDot', 'dotted', 'double', 'dashDot', 'thick', 
    # 'hair', 'mediumDashDotDot', 'thin', 'dashDotDot', 'mediumDashed', 
    # 'slantDashDot', 'medium', 'dashed'
    for fila in range(8,iterable):
        for columna in range(1,6):
            if (fila==8):
                sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                        right=Side(border_style=None, color='000000'),
                                                        top=Side(border_style='thin', color='000000'),
                                                        bottom=Side(border_style=None, color='000000'))
            if (columna==5):
                sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                        right=Side(border_style='thin', color='000000'),
                                                        top=Side(border_style=None, color='000000'),
                                                        bottom=Side(border_style=None, color='000000'))
            if (fila==(iterable)):
                sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                        right=Side(border_style=None, color='000000'),
                                                        top=Side(border_style=None, color='000000'),
                                                        bottom=Side(border_style='thin', color='000000'))
            #sheet.row_dimensions[fila].height = 80
    #fila del titular ultima_fila+1
    celda_vacia_ini=f'A{ultima_fila+1}'
    celda_vacia_fin=f'B{ultima_fila+1}'
    combinar_celdas(sheet,celda_vacia_ini,celda_vacia_fin)

    celda_titular_ini = f'C{ultima_fila+1}'
    celda_titular_fin = f'E{ultima_fila+1}'
    titular='Eduardo Alberto Berrios Villanueva'
    texto_titular = 'BCP - Titular: '+titular
    combinar_celdas(sheet,celda_titular_ini,celda_titular_fin,texto_titular)
    sheet[celda_titular_ini].alignment=Alignment(horizontal='center')

    celdaini_mensaje =f'A{ultima_fila+2}'
    celdafin_mensaje =f'E{ultima_fila+3}'
    for i in range (ultima_fila,ultima_fila+4):
        for j in range (1,6):
             sheet.cell(i,j).border = Border(left=Side(border_style='thin', color='000000'),
                                            right=Side(border_style='thin', color='000000'),
                                            top=Side(border_style='thin', color='000000'),
                                            bottom=Side(border_style='thin', color='000000'))
    mensaje_extra='Mensaje extra al pie de pagina'
    combinar_celdas(sheet,celdaini_mensaje,celdafin_mensaje,mensaje_extra)
    sheet.cell(ultima_fila+2,1).alignment = Alignment(horizontal='center',vertical='center')
    #----------------------------------------------------------------

    book.save('sample_data5.xlsx')
    book.close()

if __name__ == "__main__":
    while True:
        #tipo_moneda='S/.'
        datajson={
        "ID_Plantilla": "id1",
        "Finca_ID": "id1",
        "Seccion": [
            {
                "ID_Seccion": "id1",
                "nombre": "nombre",
                "Subsecciones": [
                    {
                        "ID_Subseccion": "id1v1",
                        "nombre": "nombre1",
                        "monto": 30.0,
                        "descripcion": "descripcion y fechas"# aca va estar PERIODO, FECHAS DE VENCIMIENTO Y EMISION
                    },
                    {
                        "ID_Subseccion": "id2v1",
                        "nombre": "nombre2",
                        "monto": 40.0,
                        "descripcion": "descripcion y fechas"
                    },
                    {
                        "ID_Subseccion": "id3v1",
                        "nombre": "nombre3",
                        "monto": 50.0,
                        "descripcion": "descripcion y fechas"
                    }
                ]
            },
            {
                "ID_Seccion": "id2",
                "nombre": "nombre",
                "Subsecciones": [
                    {
                        "ID_Subseccion": "id1_v2",
                        "nombre": "nombre2_1",
                        "monto": 60.0,
                        "descripcion": "descripcion y fechas"
                    },
                    {
                        "ID_Subseccion": "id2_v2",
                        "nombre": "nombre2_2",
                        "monto": 70.0,
                        "descripcion": "descripcion y fechas"
                    },
                    {
                        "ID_Subseccion": "id3_v2",
                        "nombre": "nombre2_3",
                        "monto": 80.0,
                        "descripcion": "descripcion y fechas"
                    }
                ]
            }
        ]
    }
        factura_excel(datajson,'$')
        ruta='C:/Users/DELL/Desktop/angular/sample_data5.xlsx'
        os.startfile(ruta,'Open')
        sys.exit()
    