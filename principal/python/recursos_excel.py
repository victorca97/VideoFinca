import pandas as pd
from openpyxl.workbook import Workbook
import openpyxl
from openpyxl import Workbook
import time
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import *
from openpyxl.utils import get_column_letter

def combinar_celdas(sheet,celda_inicial,celda_final,texto=''):
    sheet.merge_cells(f'{celda_inicial}:{celda_final}')
    if texto!='':
        sheet[celda_inicial]=texto

def formato_celdas(sheet,celda,fuente,tamaño,negrita=False,curva=False,color_texto='000000',subrayado=False):
    if (subrayado):
        tipo_subrayado='single'
    else:
        tipo_subrayado=None
    sheet[celda].font = Font(fuente,
                 size=tamaño,
                 bold=negrita,
                 italic=curva,
                 underline=tipo_subrayado,
                 color=color_texto)

def bordear_celdasv1(sheet,celda_ini,celda_final):#A1 , E8
    letra_ini_col = celda_ini[0]#string
    letra_ini_col = ord(letra_ini_col)-64#letra a numero

    letra_final_col = celda_final[0]#string
    letra_final_col = ord(letra_final_col)-64#letra a numero

    letra_ini_fil = int(celda_ini[1])#int
    letra_final_fil = int(celda_final[1])#int

    for fila in range(letra_ini_fil,letra_final_fil):
        for columna in range(letra_ini_col,letra_final_col):
            sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                    right=Side(border_style='thin', color='000000'),
                                                    top=Side(border_style='thin', color='000000'),
                                                    bottom=Side(border_style='thin', color='000000'))

def bordear_celdasv2(sheet,iterable,ultima_fila,centrado=False):#cuando tienes la posicion de las celdas en numeros
    for i in range(iterable,ultima_fila+1):
            for j in range(1,6):
                sheet.cell(i,j).border = Border(left=Side(border_style='thin', color='000000'),
                                                right=Side(border_style='thin', color='000000'),
                                                top=Side(border_style='thin', color='000000'),
                                                bottom=Side(border_style='thin', color='000000'))
                if centrado:
                    letra=get_column_letter(j)
                    celda=f'{letra}{i}'
                    sheet[celda].alignment=Alignment(horizontal='center')

def izquierda(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
    #print('izquierda')

def derecha(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                right=Side(border_style='thin', color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
    #print('derecha')

def arriba(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style='thin', color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
    #print('arriba')

def abajo(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style='thin', color='000000'))
    #print('abajo')

def bordear_lado(lado,sheet,fila,columna):
    {
        'izquierda': izquierda(sheet,fila,columna),
        'derecha':derecha(sheet,fila,columna),
        'arriba':derecha(sheet,fila,columna),
        'abajo':derecha(sheet,fila,columna)
    }.get(lado)

def getColumnName(n):#vota string
    # inicializa la string de salida como vacía
    result = ''
    while n > 0:

        # encontrar el índice de la siguiente letra y concatenar la letra
        # a la solución
 
        # aquí el índice 0 corresponde a `A`, y 25 corresponde a `Z`
        index = (n - 1) % 26
        result += chr(index + ord('A'))
        n = (n - 1) // 26
    return result[::-1]

