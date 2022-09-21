from pymongo import MongoClient
from bson.objectid import ObjectId
from bson.json_util import dumps
import sys
import xlsxwriter
from openpyxl.workbook import Workbook
import openpyxl
from openpyxl import Workbook
import time
from openpyxl.chart import ScatterChart, Reference, Series
from recursos_excel import*
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from urls import *
#Leer los datos de la BD
def leer_db(coleccion):
    bd=coleccion.find()
    for r in bd:
        #print(r['atributo']) para leer un atributo de la BD
        print(r)

def lista_json(client):
    #NOMBRE DE LA BD
    db=client["VideoFinca"]

    #NOMBRES DE LAS COLECCIONES
    collection2 = db["Finca"]
    
    #PASANDOLO A JSON
    cursor = collection2.find()
    lista_prueba=list(cursor)
    return lista_prueba,collection2

def extraer_informacion():
    json =urls()
    return json

def cantidad_finca(lista_prueba,matriz):
    i = len(lista_prueba) #sale dos
    for j in range(i):
        datos = lista_prueba[j]['_id']
        matriz.append(datos)

def cantidad_propietarios(lista_prueba,matriz):
    num_propietarios = len(lista_prueba[0]['lista_propietarios'])
    for p in range(num_propietarios):
        datos = lista_prueba[0]['lista_propietarios'][p]['_id']
        matriz.append(datos)
    #Propietarios ['123456789', '1122334455']

def principalv1(finca,collection2,tipo_moneda='S/.',n_excel=1):
    for f in finca:
        #representaciones
        plantilla2=[
        {
        "$match": {
            "_id": f'{f}'
        }
        },  
        {
        "$lookup": {
            "from": 'Plantilla',
            "localField": '_id',
            "foreignField": 'Finca',
            "as": 'Plantillas'
        }
        },  
        {
        "$lookup": {
            "from": 'Propietario',
            "localField": '_id',
            "foreignField": 'Finca',
            "as": 'Propietarios'
        }
        }
        ]
        #CREANDO LA TABLA
        #toda esta parte es fija en la tabla
        #----------------------------------------------------------------
        resultados2 =collection2.aggregate(plantilla2)
        #cursor = resultados2.find()
        list_cur = list(resultados2) #esto es una lista
        l=len(propietarios)        
        for p in propietarios:
            for i in range(l):
                propietarios=list_cur[0]['Propietarios']
                if (propietarios[i]['_id']==p):
                
                    #CREANDO EL EXCEL
                    total_monto=0
                    book=Workbook()
                    sheet= book.active
                    
                    for columna in range(1,6):
                        col_letter = get_column_letter(columna)
                        sheet.column_dimensions[col_letter].width = 20 #NO ESTA EN PIXELES
                    """propietarios_hola=list_cur[0]['Propietarios']

                    for i in propietarios_hola:
                        print(i['Correo'])
                        print(i['Departamentos'][0]['Porcentaje_Participacion'])"""

                    combinar_celdas(sheet,'B1','D1','JUNTA DE PROPIETARIOS')
                    sheet['B1'].alignment=Alignment(horizontal="center")

                    combinar_celdas(sheet,'B2','D2','EDIFICIO GALLITO DE LAS ROCAS')
                    sheet['B2'].alignment=Alignment(horizontal="center")
                    #DIRECCION
                    direccion=list_cur[0]['Direccion']
                    combinar_celdas(sheet,'B3','D3',direccion)
                    sheet['B3'].alignment=Alignment(horizontal="center")

                    combinar_celdas(sheet,'B4','D4','RECIBO POR CUOTA DE MANTENIMIENTO')
                    sheet['B4'].alignment=Alignment(horizontal="center")
                    formato_celdas(sheet,'B1','Arial',10,True,True,'000000',True)
                    combinar_celdas(sheet,'A1','A4')
                    combinar_celdas(sheet,'E1','E4')

                    sheet['A5']='Departamento:'
                    sheet['A6']='Propietario:'
                    sheet['A7']='Periodo:'

                    #DEPARTAMENTO y PORC. PARTICIPACION
                    
                    departamentos = propietarios[i]['Departamentos']
                    id_departamento = departamentos[i]['ID_Departamentos']
                    porcentaje_participacion = departamentos[i]['Porcentaje_Participacion']
                    
                    sheet['B5']=id_departamento
                    sheet['B5'].alignment=Alignment(horizontal='center')

                    sheet['B7']='SETIEMBRE 2022'
                    sheet['B7'].alignment=Alignment(horizontal='center')
                    
                    estacionamientos = propietarios[0]['Estacionamientos']

                    for j in range(len(estacionamientos)):
                        num_estacionamiento = estacionamientos[j]['Numero_Estacionamiento']
                    
                    sheet['C5']='Estacionamiento:'
                    sheet.merge_cells('B6:E6')

                    nombre = propietarios[i]['Nombre']

                    apellido = propietarios[i]['Apellido']

                    nombres_completos = nombre+' '+apellido

                    sheet['B6']=nombres_completos
                    sheet['B6'].alignment=Alignment(horizontal='center')
                    sheet['C7']='Total Porc. Part.'

                    sheet.merge_cells('D5:E5')
                    sheet['D5']=num_estacionamiento
                    sheet['D5'].alignment=Alignment(horizontal='center')

                    sheet.merge_cells('D7:E7')
                    sheet['D7']=porcentaje_participacion
                    sheet['D7'].alignment=Alignment(horizontal='center')

                    sheet['D8']='Total'
                    formato_celdas(sheet,'D8','Arial',10,False,True,'000000',True)
                    sheet['D8'].alignment=Alignment(horizontal='center')

                    sheet['E8']='Importe'
                    formato_celdas(sheet,'E8','Arial',10,False,True,'000000',True)
                    sheet['E8'].alignment=Alignment(horizontal='center')

                    for fila in range(1,8):
                        for columna in range(1,6):
                            sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                                        right=Side(border_style='thin', color='000000'),
                                                                        top=Side(border_style='thin', color='000000'),
                                                                        bottom=Side(border_style='thin', color='000000'))

                    sheet.cell(7,6).border = Border(left=Side(border_style='thin', color='000000'),
                                                    right=Side(border_style=None, color='000000'),
                                                    top=Side(border_style=None, color='000000'),
                                                    bottom=Side(border_style=None, color='000000'))
                    #----------------------------------------------------------------
                    #RECORRIENDO LOS DATOS DE CADA FINCA

                    iterable=9
                    #en esta seccion se llenan los datos de la BD
                    #----------------------------------------------------------------
                    plantillas=list_cur[0]['Plantillas']
                    #print(plantillas)
                    for j in range(len(plantillas)):
                        
                        secciones = plantillas[j]['Seccion']

                        for s in range(len(secciones)):
                            id_seccion = secciones[s]['ID_Seccion']

                            nombre = secciones[s]['nombre']

                            nombre_seccion= id_seccion+'-'+nombre
                            celda_seccion=f'A{iterable}'
                            sheet[celda_seccion]=nombre_seccion
                            
                            formato_celdas(sheet,celda_seccion,'Arial',10,True,True,'000000',True)
                            iterable=iterable+1

                            subsecciones = secciones[s]['Subsecciones']

                            for ss in range(len(subsecciones)):
                                id_subseccion = subsecciones[ss]["ID_Subseccion"]

                                nombre = subsecciones[ss]['nombre']

                                nombre_subseccion = id_subseccion+'-'+nombre
                                celda_col1=f'A{iterable}'
                                valor_col1=nombre_subseccion
                                sheet[celda_col1] = valor_col1

                                descripcion = subsecciones[ss]['descripcion']
                                celda_col2=f'B{iterable}'
                                valor_col2=descripcion
                                sheet[celda_col2] = valor_col2

                                monto = subsecciones[ss]['monto']
                                celda_col4=f'D{iterable}'
                                monto_float = "{:.2f}".format(monto)
                                if (tipo_moneda!='€'):
                                    valor_col4=tipo_moneda +' '+str(monto_float)
                                else:
                                    valor_col4=str(monto_float)+' '+tipo_moneda 
                                sheet[celda_col4] = valor_col4
                                sheet[celda_col4].alignment=Alignment(horizontal='center')

                                celda_col5=f'E{iterable}'
                                total_monto_fila=monto*porcentaje_participacion
                                total_monto=total_monto+total_monto_fila
                                total_monto_fila_float="{:.2f}".format(total_monto_fila)
                                if (tipo_moneda!='€'):
                                    valor_col5=tipo_moneda +' '+str(total_monto_fila_float)
                                else:
                                    valor_col5=str(total_monto_fila_float)+' '+tipo_moneda
                                #sheet.write(celda_col5, valor_col5)

                                sheet[celda_col5] = valor_col5
                                sheet[celda_col5].alignment=Alignment(horizontal='center')

                                celda_final_suma=celda_col5
                                iterable = iterable + 1
                    #La suma total
                    total_monto_float="{:.2f}".format(total_monto)
                    if (tipo_moneda!='€'):
                        celda_suma_expresion=tipo_moneda +' '+str(total_monto_float)
                    else:
                        celda_suma_expresion=str(total_monto_float)+' '+tipo_moneda
                    celda_total=f'A{iterable}'
                    celda_total_valor=f'E{iterable}'
                    sheet[celda_total] = 'TOTAL'
                    formato_celdas(sheet,celda_total,'Arial',10,True,False,'000000')
                    sheet[celda_total_valor]=celda_suma_expresion
                    sheet[celda_total_valor].alignment=Alignment(horizontal='center')
                    formato_celdas(sheet,celda_total_valor,'Arial',10,True,False,'000000')
                    iterable= iterable + 1
                    #----------------------------------------------------------------

                    #ultimos datos de abajo
                    #----------------------------------------------------------------
                    celda_fecha_emision=f'A{iterable}'
                    sheet[celda_fecha_emision]='Fecha de emision' #inicial
                    
                    celda_fecha_emision=f'A{iterable+1}'
                    valor_fecha_emision='01/09/2022'
                    sheet[celda_fecha_emision]=valor_fecha_emision

                    celda_fecha_vencimento=f'B{iterable}'
                    sheet[celda_fecha_vencimento]='Fecha de vencimiento'

                    celda_fecha_vencimento=f'B{iterable+1}'
                    valor_fecha_vencimento='07/07/2022'
                    sheet[celda_fecha_vencimento]=valor_fecha_vencimento

                    celda_ncuenta=f'C{iterable}'
                    sheet[celda_ncuenta]='N° Cuenta'

                    celda_CCI=f'C{iterable+1}'
                    sheet[celda_CCI]='CCI'

                    valor_ini_ncuenta=f'D{iterable}'
                    ncuenta='194-123456789'
                    sheet[valor_ini_ncuenta]=ncuenta
                    
                    valor_fin_ncuenta=f'E{iterable}'

                    valor_ini_CCI=f'D{iterable+1}'
                    cci="0021194132456789" #final
                    sheet[valor_ini_CCI]=cci

                    valor_fin_CCI=f'E{iterable+1}'
                    ultima_fila=sheet.max_row

                    for i in range(iterable,ultima_fila+1):
                        for j in range(1,6):
                            sheet.cell(i,j).border = Border(left=Side(border_style='thin', color='000000'),
                                                            right=Side(border_style='thin', color='000000'),
                                                            top=Side(border_style='thin', color='000000'),
                                                            bottom=Side(border_style='thin', color='000000'))
                            letra=get_column_letter(j)
                            celda=f'{letra}{i}'
                            sheet[celda].alignment=Alignment(horizontal='center')
                    combinar_celdas(sheet,valor_ini_ncuenta,valor_fin_ncuenta)
                    combinar_celdas(sheet,valor_ini_CCI,valor_fin_CCI)
                    
                    for fila in range(8,iterable):
                        for columna in range(1,6):
                            if (fila==8):
                                sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                                        right=Side(border_style=None, color='000000'),
                                                                        top=Side(border_style='thin', color='000000'),
                                                                        bottom=Side(border_style=None, color='000000'))
                            if (columna==5):
                                sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                                        right=Side(border_style='thin', color='000000'),
                                                                        top=Side(border_style=None, color='000000'),
                                                                        bottom=Side(border_style=None, color='000000'))
                            if (fila==(iterable)):
                                sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                                        right=Side(border_style=None, color='000000'),
                                                                        top=Side(border_style=None, color='000000'),
                                                                        bottom=Side(border_style='thin', color='000000'))
                            #sheet.row_dimensions[fila].height = 80
                    #fila del titular ultima_fila+1
                    celda_vacia_ini=f'A{ultima_fila+1}'
                    celda_vacia_fin=f'B{ultima_fila+1}'
                    combinar_celdas(sheet,celda_vacia_ini,celda_vacia_fin)

                    celda_titular_ini = f'C{ultima_fila+1}'
                    celda_titular_fin = f'E{ultima_fila+1}'
                    titular=nombres_completos
                    texto_titular = 'BCP - Titular: '+titular
                    combinar_celdas(sheet,celda_titular_ini,celda_titular_fin,texto_titular)
                    sheet[celda_titular_ini].alignment=Alignment(horizontal='center')

                    celdaini_mensaje =f'A{ultima_fila+2}'
                    celdafin_mensaje =f'E{ultima_fila+3}'
                    for i in range (ultima_fila,ultima_fila+4):
                        for j in range (1,6):
                            sheet.cell(i,j).border = Border(left=Side(border_style='thin', color='000000'),
                                                            right=Side(border_style='thin', color='000000'),
                                                            top=Side(border_style='thin', color='000000'),
                                                            bottom=Side(border_style='thin', color='000000'))
                    mensaje_extra='Mensaje extra al pie de pagina'
                    combinar_celdas(sheet,celdaini_mensaje,celdafin_mensaje,mensaje_extra)
                    sheet.cell(ultima_fila+2,1).alignment = Alignment(horizontal='center',vertical='center')
                    #----------------------------------------------------------------
                    
                    nombre_excel = f'propietarioV2_{p}_{n_excel}.xlsx'
                    book.save(nombre_excel)
                    n_excel=n_excel+1
                    book.close()

def principalv2(resultados2,propietarios,tipo_moneda='S/.',n_excel=1):

    #CREANDO LA TABLA
    #toda esta parte es fija en la tabla
    #----------------------------------------------------------------
    list_cur = list(resultados2) #esto es una lista
    l=len(propietarios)   
    prop=list_cur[0]['Propietarios']
    for i in range(l): 
        for p in propietarios:
            if (p==prop[i]['_id']):
            
                #CREANDO EL EXCEL
                total_monto=0
                book=Workbook()
                sheet= book.active
                
                for columna in range(1,6):
                    col_letter = get_column_letter(columna)
                    sheet.column_dimensions[col_letter].width = 20 #NO ESTA EN PIXELES

                combinar_celdas(sheet,'B1','D1','JUNTA DE PROPIETARIOS')
                sheet['B1'].alignment=Alignment(horizontal="center")

                combinar_celdas(sheet,'B2','D2','EDIFICIO GALLITO DE LAS ROCAS')
                sheet['B2'].alignment=Alignment(horizontal="center")
                
                #DIRECCION

                direccion=list_cur[0]['Direccion']

                combinar_celdas(sheet,'B3','D3',direccion)
                sheet['B3'].alignment=Alignment(horizontal="center")

                combinar_celdas(sheet,'B4','D4','RECIBO POR CUOTA DE MANTENIMIENTO')
                sheet['B4'].alignment=Alignment(horizontal="center")
                formato_celdas(sheet,'B1','Arial',10,True,True,'000000',True)
                combinar_celdas(sheet,'A1','A4')
                combinar_celdas(sheet,'E1','E4')

                sheet['A5']='Departamento:'
                sheet['A6']='Propietario:'
                sheet['A7']='Periodo:'

                #DEPARTAMENTO y PORC. PARTICIPACION
                
                departamentos = prop[i]['Departamentos']
                id_departamento = departamentos[0]['ID_Departamentos']
                porcentaje_participacion = departamentos[0]['Porcentaje_Participacion']
                
                sheet['B5']=id_departamento
                sheet['B5'].alignment=Alignment(horizontal='center')

                sheet['B7']='SETIEMBRE 2022'
                sheet['B7'].alignment=Alignment(horizontal='center')
                
                estacionamientos = prop[0]['Estacionamientos']

                for j in range(len(estacionamientos)):
                    num_estacionamiento = estacionamientos[j]['Numero_Estacionamiento']
                
                sheet['C5']='Estacionamiento:'
                sheet.merge_cells('B6:E6')

                nombre = prop[i]['Nombre']

                apellido = prop[i]['Apellido']

                nombres_completos = nombre+' '+apellido

                sheet['B6']=nombres_completos
                sheet['B6'].alignment=Alignment(horizontal='center')
                sheet['C7']='Total Porc. Part.'

                sheet.merge_cells('D5:E5')
                sheet['D5']=num_estacionamiento
                sheet['D5'].alignment=Alignment(horizontal='center')

                sheet.merge_cells('D7:E7')
                sheet['D7']=porcentaje_participacion
                sheet['D7'].alignment=Alignment(horizontal='center')

                sheet['D8']='Total'
                formato_celdas(sheet,'D8','Arial',10,False,True,'000000',True)
                sheet['D8'].alignment=Alignment(horizontal='center')

                sheet['E8']='Importe'
                formato_celdas(sheet,'E8','Arial',10,False,True,'000000',True)
                sheet['E8'].alignment=Alignment(horizontal='center')

                for fila in range(1,8):
                    for columna in range(1,6):
                        sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                                    right=Side(border_style='thin', color='000000'),
                                                                    top=Side(border_style='thin', color='000000'),
                                                                    bottom=Side(border_style='thin', color='000000'))

                sheet.cell(7,6).border = Border(left=Side(border_style='thin', color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
                #----------------------------------------------------------------
                #RECORRIENDO LOS DATOS DE CADA FINCA
                iterable=9
                #en esta seccion se llenan los datos de la BD
                #----------------------------------------------------------------
                plantillas=list_cur[0]['Plantillas']
                #print(plantillas)
                for j in range(len(plantillas)):
                    
                    secciones = plantillas[j]['Seccion']

                    for s in range(len(secciones)):
                        id_seccion = secciones[s]['ID_Seccion']

                        nombre = secciones[s]['nombre']

                        nombre_seccion= id_seccion+'-'+nombre
                        celda_seccion=f'A{iterable}'
                        sheet[celda_seccion]=nombre_seccion
                        
                        formato_celdas(sheet,celda_seccion,'Arial',10,True,True,'000000',True)
                        iterable=iterable+1

                        subsecciones = secciones[s]['Subsecciones']

                        for ss in range(len(subsecciones)):
                            id_subseccion = subsecciones[ss]["ID_Subseccion"]

                            nombre = subsecciones[ss]['nombre']

                            nombre_subseccion = id_subseccion+'-'+nombre
                            celda_col1=f'A{iterable}'
                            valor_col1=nombre_subseccion
                            sheet[celda_col1] = valor_col1

                            descripcion = subsecciones[ss]['descripcion']
                            celda_col2=f'B{iterable}'
                            valor_col2=descripcion
                            sheet[celda_col2] = valor_col2

                            monto = subsecciones[ss]['monto']
                            celda_col4=f'D{iterable}'
                            monto_float = "{:.2f}".format(monto)
                            if (tipo_moneda!='€'):
                                valor_col4=tipo_moneda +' '+str(monto_float)
                            else:
                                valor_col4=str(monto_float)+' '+tipo_moneda 
                            sheet[celda_col4] = valor_col4
                            sheet[celda_col4].alignment=Alignment(horizontal='center')

                            celda_col5=f'E{iterable}'
                            total_monto_fila=monto*porcentaje_participacion
                            total_monto=total_monto+total_monto_fila
                            total_monto_fila_float="{:.2f}".format(total_monto_fila)
                            if (tipo_moneda!='€'):
                                valor_col5=tipo_moneda +' '+str(total_monto_fila_float)
                            else:
                                valor_col5=str(total_monto_fila_float)+' '+tipo_moneda
                            #sheet.write(celda_col5, valor_col5)

                            sheet[celda_col5] = valor_col5
                            sheet[celda_col5].alignment=Alignment(horizontal='center')

                            celda_final_suma=celda_col5
                            iterable = iterable + 1
                #La suma total
                total_monto_float="{:.2f}".format(total_monto)
                if (tipo_moneda!='€'):
                    celda_suma_expresion=tipo_moneda +' '+str(total_monto_float)
                else:
                    celda_suma_expresion=str(total_monto_float)+' '+tipo_moneda
                celda_total=f'A{iterable}'
                celda_total_valor=f'E{iterable}'
                sheet[celda_total] = 'TOTAL'
                formato_celdas(sheet,celda_total,'Arial',10,True,False,'000000')
                sheet[celda_total_valor]=celda_suma_expresion
                sheet[celda_total_valor].alignment=Alignment(horizontal='center')
                formato_celdas(sheet,celda_total_valor,'Arial',10,True,False,'000000')
                iterable= iterable + 1
                #----------------------------------------------------------------

                #ultimos datos de abajo
                #----------------------------------------------------------------
                celda_fecha_emision=f'A{iterable}'
                sheet[celda_fecha_emision]='Fecha de emision' #inicial
                
                celda_fecha_emision=f'A{iterable+1}'
                valor_fecha_emision='01/09/2022'
                sheet[celda_fecha_emision]=valor_fecha_emision

                celda_fecha_vencimento=f'B{iterable}'
                sheet[celda_fecha_vencimento]='Fecha de vencimiento'

                celda_fecha_vencimento=f'B{iterable+1}'
                valor_fecha_vencimento='07/07/2022'
                sheet[celda_fecha_vencimento]=valor_fecha_vencimento

                celda_ncuenta=f'C{iterable}'
                sheet[celda_ncuenta]='N° Cuenta'

                celda_CCI=f'C{iterable+1}'
                sheet[celda_CCI]='CCI'

                valor_ini_ncuenta=f'D{iterable}'
                ncuenta='194-123456789'
                sheet[valor_ini_ncuenta]=ncuenta
                
                valor_fin_ncuenta=f'E{iterable}'

                valor_ini_CCI=f'D{iterable+1}'
                cci="0021194132456789" #final
                sheet[valor_ini_CCI]=cci

                valor_fin_CCI=f'E{iterable+1}'
                ultima_fila=sheet.max_row

                for i in range(iterable,ultima_fila+1):
                    for j in range(1,6):
                        sheet.cell(i,j).border = Border(left=Side(border_style='thin', color='000000'),
                                                        right=Side(border_style='thin', color='000000'),
                                                        top=Side(border_style='thin', color='000000'),
                                                        bottom=Side(border_style='thin', color='000000'))
                        letra=get_column_letter(j)
                        celda=f'{letra}{i}'
                        sheet[celda].alignment=Alignment(horizontal='center')
                combinar_celdas(sheet,valor_ini_ncuenta,valor_fin_ncuenta)
                combinar_celdas(sheet,valor_ini_CCI,valor_fin_CCI)
                
                for fila in range(8,iterable):
                    for columna in range(1,6):
                        if (fila==8):
                            sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                                    right=Side(border_style=None, color='000000'),
                                                                    top=Side(border_style='thin', color='000000'),
                                                                    bottom=Side(border_style=None, color='000000'))
                        if (columna==5):
                            sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                                    right=Side(border_style='thin', color='000000'),
                                                                    top=Side(border_style=None, color='000000'),
                                                                    bottom=Side(border_style=None, color='000000'))
                        if (fila==(iterable)):
                            sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                                    right=Side(border_style=None, color='000000'),
                                                                    top=Side(border_style=None, color='000000'),
                                                                    bottom=Side(border_style='thin', color='000000'))
                        #sheet.row_dimensions[fila].height = 80
                #fila del titular ultima_fila+1
                celda_vacia_ini=f'A{ultima_fila+1}'
                celda_vacia_fin=f'B{ultima_fila+1}'
                combinar_celdas(sheet,celda_vacia_ini,celda_vacia_fin)

                celda_titular_ini = f'C{ultima_fila+1}'
                celda_titular_fin = f'E{ultima_fila+1}'
                titular=nombres_completos
                texto_titular = 'BCP - Titular: '+titular
                combinar_celdas(sheet,celda_titular_ini,celda_titular_fin,texto_titular)
                sheet[celda_titular_ini].alignment=Alignment(horizontal='center')

                celdaini_mensaje =f'A{ultima_fila+2}'
                celdafin_mensaje =f'E{ultima_fila+3}'
                for i in range (ultima_fila,ultima_fila+4):
                    for j in range (1,6):
                        sheet.cell(i,j).border = Border(left=Side(border_style='thin', color='000000'),
                                                        right=Side(border_style='thin', color='000000'),
                                                        top=Side(border_style='thin', color='000000'),
                                                        bottom=Side(border_style='thin', color='000000'))
                mensaje_extra='Mensaje extra al pie de pagina'
                combinar_celdas(sheet,celdaini_mensaje,celdafin_mensaje,mensaje_extra)
                sheet.cell(ultima_fila+2,1).alignment = Alignment(horizontal='center',vertical='center')
                #----------------------------------------------------------------
                ruta_excel='C:/Users/DELL/Desktop/angular/mongodb/principal/excels/pruebas_reales'
                nombre_excel = f'/propietarioV3_{p}_{n_excel}.xlsx'
                excel_guardar = ruta_excel+nombre_excel
                book.save(excel_guardar)
                n_excel=n_excel+1
                print('Excel creado')
                book.close()  
                break