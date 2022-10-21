import base64
import os
from tkinter import E
import openpyxl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
import xlwings as xw
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image
import pathlib

def combinar_celdas(sheet,celda_inicial,celda_final,texto=''):
    sheet.merge_cells(f'{celda_inicial}:{celda_final}')
    if texto!='':
        sheet[celda_inicial]=texto

def formato_celdas(sheet,celda,fuente,tamaño,negrita=False,curva=False,color_texto='000000',subrayado=False):
    if (subrayado):
        tipo_subrayado='single'
    else:
        tipo_subrayado=None
    sheet[celda].font = Font(fuente,
                 size=tamaño,
                 bold=negrita,
                 italic=curva,
                 underline=tipo_subrayado,
                 color=color_texto)

def bordear_celdasv1(sheet,celda_ini,celda_final):#A1 , E8
    letra_ini_col = celda_ini[0]#string
    letra_ini_col = ord(letra_ini_col)-64#letra a numero

    letra_final_col = celda_final[0]#string
    letra_final_col = ord(letra_final_col)-64#letra a numero

    letra_ini_fil = int(celda_ini[1])#int
    letra_final_fil = int(celda_final[1])#int

    for fila in range(letra_ini_fil,letra_final_fil):
        for columna in range(letra_ini_col,letra_final_col):
            sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                    right=Side(border_style='thin', color='000000'),
                                                    top=Side(border_style='thin', color='000000'),
                                                    bottom=Side(border_style='thin', color='000000'))

def bordear_celdasv2(sheet,iterable,ultima_fila,centrado=False):#cuando tienes la posicion de las celdas en numeros
    for i in range(iterable,ultima_fila+1):
            for j in range(1,6):
                sheet.cell(i,j).border = Border(left=Side(border_style='thin', color='000000'),
                                                right=Side(border_style='thin', color='000000'),
                                                top=Side(border_style='thin', color='000000'),
                                                bottom=Side(border_style='thin', color='000000'))
                if centrado:
                    letra=get_column_letter(j)
                    celda=f'{letra}{i}'
                    sheet[celda].alignment=Alignment(horizontal='center')

def izquierda(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
    #print('izquierda')

def derecha(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                right=Side(border_style='thin', color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
    #print('derecha')

def arriba(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style='thin', color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
    #print('arriba')

def abajo(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style='thin', color='000000'))
    #print('abajo')

def bordear_lado(lado,sheet,fila,columna):
    {
        'izquierda': izquierda(sheet,fila,columna),
        'derecha':derecha(sheet,fila,columna),
        'arriba':derecha(sheet,fila,columna),
        'abajo':derecha(sheet,fila,columna)
    }.get(lado)

def getColumnName(n):#vota string
    # inicializa la string de salida como vacía
    result = ''
    while n > 0:

        # encontrar el índice de la siguiente letra y concatenar la letra
        # a la solución
 
        # aquí el índice 0 corresponde a `A`, y 25 corresponde a `Z`
        index = (n - 1) % 26
        result += chr(index + ord('A'))
        n = (n - 1) // 26
    return result[::-1]

def ancho_col(sheet):
    lista_ancho_columnas=[17,18,14,10,17]
    anchos=0
    for columna in range(1,6):
        col_letter = get_column_letter(columna)
        medida=lista_ancho_columnas[anchos]
        sheet.column_dimensions[col_letter].width = medida #NO ESTA EN PIXELES
        anchos=anchos+1

def convertir_pdf(varbuffer,cantidad_propietarios,path): 
    #version original
    """list_dir = [] 
    list_dir = os.listdir(path)
    for file in list_dir: 
        if file.endswith(tipo): # eg: '.txt' 
            #print(file) nombre de los excels
            #file = propietario_eduardo berrios.xlsx
            nombre_archivo = file #nombre del archivo como tal (con la extension)
            auxiliar1 = nombre_archivo.split('.')
            nombre_sin_extension = auxiliar1[0]#le quitamos la extension .xlsx 
            #nombre_sin_extension = propietario_eduardo berrios
            auxiliar2 = nombre_sin_extension.split('_')
            nombre_propietario = auxiliar2[1]
            #nombre_propietario = eduardo berrios
            ruta_excel = path+'/'+nombre_archivo
            pdf_codificado = convertir_a_pdf(ruta_excel,nombre_sin_extension)"""
    #version adaptada
    for i in range(cantidad_propietarios):
        if (varbuffer[i][0]!=1):
            nombre_archivo = f'propietario_{varbuffer[i][1]}'
            ruta_excel = path+'/'+nombre_archivo+'.xlsx'
            pdf_codificado = convertir_a_pdf(ruta_excel,nombre_archivo)
            varbuffer[i][3]= pdf_codificado
    return varbuffer

def convertir_a_pdf(ruta_excel,nombre_archivo):#FALLA SI EL PDF YA EXISTE
    excel_app = xw.App(visible=False)
    print('Iniciando ...')
    try:
        # Initialize new excel workbook
        #book = load_workbook(ruta_excel+'/'+nombre_archivo+'.xlsx')
        #book = xw.Book(ruta_excel)#RUTA
        excel_book = excel_app.books.open(ruta_excel)

        ruta = str(pathlib.Path().absolute())
        a=ruta.replace('\\','/')
        x = a.find("/",-1,0)
        len_x=len(a)
        i=2
        while(x==-1):
            x = a.find("/",len_x-i,len_x-1)
            i=i+1
        ruta_pdf = a[0:x+1]

        # ruta_pdf='C:/Users/DELL/Desktop/angular/mongodb/principal/excels/pruebas/'

        nombre_pdf = nombre_archivo+'.pdf'
        pdf_path = ruta_pdf+nombre_pdf
        # Save excel workbook to pdf file
        print(f"Saving workbook as '{pdf_path}' ...")
        excel_book.api.ExportAsFixedFormat(0, pdf_path)
        print('conversion exitosa')
        excel_book.close()
        excel_app.quit()
        #convertir_base64 el PDF
        pdf_codificado = convertir_base64(pdf_path)
        return pdf_codificado
    except Exception as e:
        excel_book.close()
        excel_app.quit()
        print(e)

def poner_imagenes(sheet):
    #insertando la imagen
        #logo = openpyxl.drawing.image.Image('C:/Users/DELL/Desktop/angular/mongodb/principal/LOGOVIDEOFINCA_ORIGINAL.png')
        logo_pil = Image.open('../excels/recursos/LOGOVIDEOFINCA_ORIGINAL.png')
        finca_pil = Image.open('../excels/recursos/FINCA.jpg')
        
        #establecer dimensiones
        proporcion = 3
        alto = 40
        ancho = int(proporcion * alto)

        #editar la imagen
        logo_pil = logo_pil.resize((ancho,alto+5))#140 80 solo acepta enteros
        logo_pil.save('../excels/recursos/LOGOVIDEOFINCA_MODIFICADO.png')

        finca_pil = finca_pil.resize((ancho+3,alto))
        finca_pil.save('../excels/recursos/FINCA_MODIFICADO.jpg')
    
        logo = openpyxl.drawing.image.Image('../excels/recursos/LOGOVIDEOFINCA_MODIFICADO.png')
        
        p2e = pixels_to_EMU
        c2e = cm_to_EMU
        h, w = logo.height, logo.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        # Calculated number of cells width or height from cm into EMUs
        cellh = lambda x: c2e((x * 49.77)/99)
        cellw = lambda x: c2e((x * (18.65-1.71))/10)

        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
        # Also offset by half a column.
        column = 0
        coloffset = cellw(0.05)
        row = 1
        rowoffset = cellh(0.05)

        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
        logo.anchor = OneCellAnchor(_from=marker, ext=size)

        sheet.add_image(logo)
        

        finca = openpyxl.drawing.image.Image('../excels/recursos/FINCA_MODIFICADO.jpg')
        p2e = pixels_to_EMU
        c2e = cm_to_EMU
        h, w = finca.height, finca.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        # Calculated number of cells width or height from cm into EMUs
        cellh = lambda x: c2e((x * 49.77)/99)
        cellw = lambda x: c2e((x * (18.65-1.71))/10)

        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
        # Also offset by half a column.
        column = 4
        coloffset = cellw(0.05)
        row = 1
        rowoffset = cellh(0.05)

        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
        finca.anchor = OneCellAnchor(_from=marker, ext=size)

        sheet.add_image(finca) 

def ancho_columnas_parametros(sheet):
    lista_ancho_columnas=[17,20,15,20,17]
    anchos=0
    for columna in range(1,6):
        col_letter = get_column_letter(columna)
        medida=lista_ancho_columnas[anchos]
        sheet.column_dimensions[col_letter].width = medida #NO ESTA EN PIXELES
        anchos=anchos+1

def convertir_base64(ruta):# ruta: C:/Users/DELL/Desktop/excel.xlsx
    with open(ruta, 'rb') as binary_file:
        binary_file_data = binary_file.read()
        base64_encoded_data = base64.b64encode(binary_file_data)
        base64_message = base64_encoded_data.decode('utf-8')
        #print(base64_message)
    return base64_message

def decodificar_base64(codificado,nombre_archivo,extension):
    base64_img_bytes = codificado.encode('utf-8')
    with open(f'C:/Users/DELL/Desktop/{nombre_archivo}_decodificado.{extension}', 'wb') as file_to_save:
        decoded_image_data = base64.decodebytes(base64_img_bytes)
        file_to_save.write(decoded_image_data)

#PRUEBAS        
def prueba():
    codificado = convertir_base64('C:/Users/DELL/Desktop/angular/mongodb/principal/excels/pruebas/propietarioV4_123456789_2.xlsx')
    decodificar_base64(codificado,'propietarioV4_123456789_2','xlsx')